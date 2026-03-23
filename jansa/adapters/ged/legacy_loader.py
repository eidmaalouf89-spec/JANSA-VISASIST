"""Legacy document loader — extracts ancien flags from GrandFichier.

The GrandFichier is a team-manipulated GED export that contains an 'ANCIEN'
column marking legacy documents (value=1) from the previous system.

Documents flagged ancien=1 must be excluded from the operational workflow.
This module extracts those flags and provides a cross-reference set that
NM5 uses to exclude legacy docs from the active dataset.

Join key: (numero, indice) — the 6-digit N° Doc padded with leading zeros
matched against the pipeline's `numero` column, plus the revision INDICE.
"""

import warnings
from pathlib import Path
from typing import Optional

import pandas as pd

from jansa.adapters.ged.logging import log_event


def load_ancien_flags(grandfichier_path: str) -> set:
    """Load ancien=1 flags from GrandFichier.

    Reads all sheets, finds header row containing 'ANCIEN' column,
    and collects (numero_padded, indice) tuples for rows where ancien=1.

    Args:
        grandfichier_path: Path to GrandFichier_1.xlsx

    Returns:
        Set of (numero: str, indice: str) tuples where ancien=1.
        numero is zero-padded to 6 digits. indice is uppercase.
        Empty set if file not found or no ancien flags.
    """
    path = Path(grandfichier_path)
    if not path.exists():
        log_event(
            None, 'LEGACY', 'WARNING', 'GRANDFICHIER_NOT_FOUND',
            f'GrandFichier not found at {grandfichier_path}',
        )
        return set()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        log_event(
            None, 'LEGACY', 'WARNING', 'GRANDFICHIER_LOAD_ERROR',
            f'Failed to load GrandFichier: {e}',
        )
        return set()

    ancien_set = set()
    total_ancien = 0
    sheets_processed = 0

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Locate header row by finding ANCIEN, DOCUMENT, N° Doc, IND columns
            ancien_col = None
            doc_col = None
            ndoc_col = None
            ind_col = None
            header_row_idx = None

            for i, row in enumerate(ws.iter_rows(max_row=15, values_only=False)):
                for j, cell in enumerate(row):
                    if cell.value is None:
                        continue
                    val = str(cell.value).strip()
                    val_upper = val.upper()
                    if val_upper == 'ANCIEN':
                        ancien_col = j
                    if val_upper == 'DOCUMENT':
                        doc_col = j
                    if val == 'N° Doc' or val_upper == 'N° DOC':
                        ndoc_col = j
                    if val == 'IND' or val_upper == 'IND':
                        ind_col = j

                if ancien_col is not None and doc_col is not None:
                    header_row_idx = i
                    break

            if header_row_idx is None or ancien_col is None:
                continue

            sheets_processed += 1
            sheet_ancien = 0

            for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
                # Skip empty rows
                doc_val = row[doc_col] if doc_col < len(row) else None
                if doc_val is None or str(doc_val).strip() == '':
                    continue

                # Check ancien flag
                ancien_val = row[ancien_col] if ancien_col < len(row) else None
                if ancien_val != 1 and str(ancien_val).strip() != '1':
                    continue

                # Extract N° Doc and IND
                ndoc_val = row[ndoc_col] if ndoc_col is not None and ndoc_col < len(row) else None
                ind_val = row[ind_col] if ind_col is not None and ind_col < len(row) else None

                if ndoc_val is None:
                    continue

                # Normalize N° Doc to 6-digit zero-padded string
                try:
                    if isinstance(ndoc_val, (int, float)):
                        ndoc_str = str(int(ndoc_val)).zfill(6)
                    else:
                        ndoc_str = str(ndoc_val).strip().zfill(6)
                except (ValueError, TypeError):
                    continue

                # Normalize indice to uppercase
                ind_str = str(ind_val).strip().upper() if ind_val is not None else ''

                ancien_set.add((ndoc_str, ind_str))
                sheet_ancien += 1
                total_ancien += 1

            if sheet_ancien > 0:
                log_event(
                    None, 'LEGACY', 'INFO', 'SHEET_ANCIEN_COUNT',
                    f'Sheet "{sheet_name}": {sheet_ancien} ancien rows',
                )

    wb.close()

    log_event(
        None, 'LEGACY', 'INFO', 'ANCIEN_LOAD_COMPLETE',
        f'Loaded {total_ancien} ancien flags from {sheets_processed} sheets, '
        f'{len(ancien_set)} unique (numero, indice) pairs',
    )

    return ancien_set


def flag_legacy_docs(
    doc_level: pd.DataFrame,
    ancien_set: set,
    numero_col: str = 'numero',
    indice_col: str = 'indice',
) -> pd.DataFrame:
    """Flag legacy documents in doc_level DataFrame.

    Adds columns:
        - is_legacy (bool): True if (numero, indice) found in ancien_set
        - exclusion_reason (str): 'LEGACY_ANCIEN' if is_legacy, else None

    Args:
        doc_level: DataFrame with one row per doc_id (from NM5 step 1)
        ancien_set: Set of (numero, indice) tuples from load_ancien_flags()
        numero_col: Column name containing the document numero
        indice_col: Column name containing the revision indice

    Returns:
        doc_level with is_legacy and exclusion_reason columns added.
    """
    if not ancien_set:
        doc_level['is_legacy'] = False
        doc_level['exclusion_reason'] = None
        return doc_level

    def _check_legacy(row) -> bool:
        numero = str(row[numero_col]).strip().zfill(6) if pd.notna(row[numero_col]) else ''
        indice = str(row[indice_col]).strip().upper() if pd.notna(row[indice_col]) else ''
        return (numero, indice) in ancien_set

    doc_level['is_legacy'] = doc_level.apply(_check_legacy, axis=1)
    doc_level['exclusion_reason'] = doc_level['is_legacy'].map(
        lambda x: 'LEGACY_ANCIEN' if x else None
    )

    legacy_count = doc_level['is_legacy'].sum()
    if legacy_count > 0:
        log_event(
            None, 'LEGACY', 'INFO', 'LEGACY_FLAGGED',
            f'{legacy_count} documents flagged as legacy (ancien=1)',
        )

    return doc_level
